#!/usr/bin/env python3
"""Acquire official public disaster-support sources used by the web product.

Sources:
- Satu Data Sumbar CKAN datasets produced by BPBD Provinsi Sumatera Barat / Pusdalops.
- Badan Informasi Geospasial (BIG) ArcGIS REST administrative boundaries.

The script deliberately does not reinterpret impact semantics. XLSX worksheets are
materialized as source-native CSV grids so the actual headers/shape can be reviewed
before promotion into canonical observations. Raw XLSX bytes are not committed;
their SHA-256 values and transport metadata are frozen in manifests.
"""

from __future__ import annotations

import csv
import hashlib
import json
import re
import urllib.parse
import urllib.request
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
BPBD_OUTPUT = ROOT / "data/processed/bpbd/sumbar_open_data/2024"
GEO_OUTPUT = ROOT / "data/processed/geography/sumbar-big-kabkota.geojson"
MANIFEST_DIR = ROOT / "data/manifests"
BPBD_MANIFEST = MANIFEST_DIR / "sumbar_bpbd_open_data_2024.json"
GEO_MANIFEST = MANIFEST_DIR / "sumbar_big_kabkota_boundary.json"

CKAN_BASES = (
    "https://data.sumbarprov.go.id/api/3/action",
    "https://data.sumbarprov.go.id/id/api/3/action",
)

BPBD_PACKAGES = (
    "jumlah-korban-per-kabkota-2024",
    "jumlah-kejadian-bencana-per-kabkota-2024",
    "dampak-bencana-terhadap-pemukiman-per-kabkota-2024",
    "dampak-bencana-terhadap-fasilitas-umum-per-kabkota-2024",
)

# June 2026 BIG national administrative-area service. Use AR (area/polygon), not
# LN (boundary line), because the public product needs district polygons.
BIG_LAYER = "https://geoservices.big.go.id/rbi/rest/services/BATASWILAYAH/BATAS_KABKOTA_AR/MapServer/0"
BIG_QUERY = BIG_LAYER + "/query"


def fetch_bytes(url: str, timeout: int = 60) -> bytes:
    request = urllib.request.Request(
        url,
        headers={"User-Agent": "RanahObservatory/1.0 (+https://github.com/nabilrn/ranah-observatory)"},
    )
    with urllib.request.urlopen(request, timeout=timeout) as response:
        return response.read()


def fetch_json(url: str, timeout: int = 60) -> dict:
    payload = json.loads(fetch_bytes(url, timeout=timeout).decode("utf-8"))
    if isinstance(payload, dict) and payload.get("error"):
        raise RuntimeError(f"remote API error from {url}: {payload['error']}")
    return payload


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def sha256_path(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def ckan_action(action: str, **params: str) -> dict:
    query = urllib.parse.urlencode(params)
    errors: list[str] = []
    for base in CKAN_BASES:
        url = f"{base}/{action}?{query}"
        try:
            payload = fetch_json(url)
            if payload.get("success") is not True:
                raise RuntimeError(f"CKAN success=false: {payload}")
            return payload["result"]
        except Exception as exc:
            errors.append(f"{url}: {exc}")
    raise RuntimeError("all official CKAN endpoints failed:\n" + "\n".join(errors))


def safe_component(value: str) -> str:
    value = re.sub(r"[^a-zA-Z0-9._-]+", "-", value.strip()).strip("-")
    return value or "sheet"


def cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value)


def materialize_workbook(package_slug: str, package: dict) -> dict:
    resources = [
        item
        for item in package.get("resources", [])
        if str(item.get("format", "")).strip().upper() in {"XLSX", "XLS"}
        and str(item.get("url", "")).startswith("https://data.sumbarprov.go.id/")
    ]
    if not resources:
        raise RuntimeError(f"{package_slug}: no official XLSX resource found")

    package_title = str(package.get("title", "")).casefold()
    resource = sorted(
        resources,
        key=lambda item: (
            package_title not in str(item.get("name", "")).casefold(),
            str(item.get("name", "")),
        ),
    )[0]

    download_url = str(resource["url"])
    workbook_bytes = fetch_bytes(download_url, timeout=120)
    digest = sha256_bytes(workbook_bytes)

    workbook = load_workbook(BytesIO(workbook_bytes), read_only=True, data_only=True)
    package_dir = BPBD_OUTPUT / package_slug
    package_dir.mkdir(parents=True, exist_ok=True)

    for old_csv in package_dir.glob("*.csv"):
        old_csv.unlink()

    sheet_records: list[dict] = []
    for index, worksheet in enumerate(workbook.worksheets, start=1):
        filename = f"{index:02d}-{safe_component(worksheet.title)}.csv"
        output = package_dir / filename
        row_count = 0
        max_columns = 0
        nonempty_rows = 0
        with output.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            for row in worksheet.iter_rows(values_only=True):
                values = [cell_text(value) for value in row]
                while values and values[-1] == "":
                    values.pop()
                writer.writerow(values)
                row_count += 1
                max_columns = max(max_columns, len(values))
                if any(value != "" for value in values):
                    nonempty_rows += 1
        sheet_records.append(
            {
                "sheet_index": index,
                "sheet_title": worksheet.title,
                "path": output.relative_to(ROOT).as_posix(),
                "sha256": sha256_path(output),
                "row_count": row_count,
                "nonempty_rows": nonempty_rows,
                "max_columns": max_columns,
            }
        )

    return {
        "package_slug": package_slug,
        "package_id": package.get("id"),
        "title": package.get("title"),
        "organization": (package.get("organization") or {}).get("title"),
        "author": package.get("author"),
        "metadata_created": package.get("metadata_created"),
        "metadata_modified": package.get("metadata_modified"),
        "resource_id": resource.get("id"),
        "resource_name": resource.get("name"),
        "resource_format": resource.get("format"),
        "resource_url": download_url,
        "resource_last_modified": resource.get("last_modified"),
        "download_sha256": digest,
        "download_size_bytes": len(workbook_bytes),
        "worksheets": sheet_records,
    }


def lower_property(properties: dict, key: str):
    key = key.casefold()
    for candidate, value in properties.items():
        if str(candidate).casefold() == key:
            return value
    return None


def polygon_parts(geometry: dict | None) -> list:
    if not geometry:
        return []
    geometry_type = geometry.get("type")
    coordinates = geometry.get("coordinates") or []
    if geometry_type == "Polygon":
        return [coordinates]
    if geometry_type == "MultiPolygon":
        return list(coordinates)
    raise RuntimeError(f"unexpected BIG geometry type: {geometry_type!r}")


def first_nonempty(values: list):
    for value in values:
        if value not in (None, ""):
            return value
    return None


def acquire_big_boundary() -> dict:
    layer_metadata = fetch_json(BIG_LAYER + "?f=pjson")
    if layer_metadata.get("geometryType") != "esriGeometryPolygon":
        raise RuntimeError(f"BIG kab/kota source is not polygon geometry: {layer_metadata.get('geometryType')}")

    params = {
        "where": "WADMPR='Sumatera Barat'",
        "outFields": "*",
        "returnGeometry": "true",
        "outSR": "4326",
        "f": "geojson",
    }
    query_url = BIG_QUERY + "?" + urllib.parse.urlencode(params)
    payload = fetch_json(query_url, timeout=120)
    source_features = payload.get("features") or []
    if not source_features:
        raise RuntimeError("BIG Sumatera Barat kab/kota query returned no features")

    grouped: dict[str, dict] = {}
    for feature in source_features:
        properties = feature.get("properties") or {}
        province = lower_property(properties, "wadmpr")
        name = lower_property(properties, "wadmkk") or lower_property(properties, "namobj")
        if str(province).strip().casefold() != "sumatera barat":
            raise RuntimeError(f"BIG query returned non-Sumatera Barat feature: {province!r} {name!r}")
        if not name:
            raise RuntimeError("BIG feature missing kabupaten/kota name")
        name = str(name).strip()
        key = name.casefold()
        record = grouped.setdefault(
            key,
            {
                "name": name,
                "polygon_parts": [],
                "kdbbps": [],
                "kdcbps": [],
                "kdpkab": [],
                "source_objectids": [],
            },
        )
        record["polygon_parts"].extend(polygon_parts(feature.get("geometry")))
        record["kdbbps"].append(lower_property(properties, "kdbbps"))
        record["kdcbps"].append(lower_property(properties, "kdcbps"))
        record["kdpkab"].append(lower_property(properties, "kdpkab"))
        record["source_objectids"].append(
            lower_property(properties, "objectid") or lower_property(properties, "objectid_1")
        )

    if len(grouped) != 19:
        names = sorted(record["name"] for record in grouped.values())
        raise RuntimeError(
            f"BIG Sumatera Barat query produced {len(source_features)} source features but {len(grouped)} unique kab/kota; "
            f"expected 19. names={names}"
        )

    normalized_features = []
    for record in grouped.values():
        parts = record["polygon_parts"]
        if not parts:
            raise RuntimeError(f"BIG feature has no polygon geometry: {record['name']}")
        normalized_features.append(
            {
                "type": "Feature",
                "geometry": {"type": "MultiPolygon", "coordinates": parts},
                "properties": {
                    "name": record["name"],
                    "province": "Sumatera Barat",
                    "kdbbps": first_nonempty(record["kdbbps"]),
                    "kdcbps": first_nonempty(record["kdcbps"]),
                    "kdpkab": first_nonempty(record["kdpkab"]),
                    "source_objectids": [value for value in record["source_objectids"] if value not in (None, "")],
                    "source_feature_count": len(record["source_objectids"]),
                },
            }
        )

    normalized_features.sort(key=lambda item: item["properties"]["name"].casefold())
    output_payload = {
        "type": "FeatureCollection",
        "name": "Sumatera Barat Kabupaten/Kota",
        "features": normalized_features,
    }
    GEO_OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    GEO_OUTPUT.write_text(json.dumps(output_payload, ensure_ascii=False, separators=(",", ":")) + "\n", encoding="utf-8")

    return {
        "source_organization": "Badan Informasi Geospasial",
        "layer_url": BIG_LAYER,
        "query_url": query_url,
        "service_item_id": layer_metadata.get("serviceItemId"),
        "layer_name": layer_metadata.get("name"),
        "geometry_type": layer_metadata.get("geometryType"),
        "spatial_reference": (layer_metadata.get("extent") or {}).get("spatialReference"),
        "supported_query_formats": layer_metadata.get("supportedQueryFormats"),
        "source_feature_count": len(source_features),
        "feature_count": len(normalized_features),
        "district_names": [item["properties"]["name"] for item in normalized_features],
        "output_path": GEO_OUTPUT.relative_to(ROOT).as_posix(),
        "output_sha256": sha256_path(GEO_OUTPUT),
    }


def main() -> None:
    retrieved_at = datetime.now(timezone.utc).isoformat()
    BPBD_OUTPUT.mkdir(parents=True, exist_ok=True)
    MANIFEST_DIR.mkdir(parents=True, exist_ok=True)

    bpbd_records = []
    for slug in BPBD_PACKAGES:
        package = ckan_action("package_show", id=slug)
        record = materialize_workbook(slug, package)
        bpbd_records.append(record)
        print(json.dumps({"bpbd_package": slug, "worksheets": len(record["worksheets"]), "sha256": record["download_sha256"]}))

    BPBD_MANIFEST.write_text(
        json.dumps(
            {
                "schema": "ranah-observatory/sumbar-bpbd-open-data-acquisition/v1",
                "retrieved_at": retrieved_at,
                "source": "https://data.sumbarprov.go.id/",
                "producer": "BPBD Provinsi Sumatera Barat",
                "source_data": "Pusdalops BPBD Sumatera Barat",
                "year": 2024,
                "promotion_state": "source_native_review_required",
                "missing_values_inferred": False,
                "packages": bpbd_records,
            },
            ensure_ascii=False,
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )

    geo_record = acquire_big_boundary()
    GEO_MANIFEST.write_text(
        json.dumps(
            {
                "schema": "ranah-observatory/big-kabkota-boundary-acquisition/v1",
                "retrieved_at": retrieved_at,
                **geo_record,
            },
            ensure_ascii=False,
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )

    print(
        json.dumps(
            {
                "bpbd_manifest": BPBD_MANIFEST.relative_to(ROOT).as_posix(),
                "bpbd_package_count": len(bpbd_records),
                "geo_manifest": GEO_MANIFEST.relative_to(ROOT).as_posix(),
                "geo_source_feature_count": geo_record["source_feature_count"],
                "geo_feature_count": geo_record["feature_count"],
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
