#!/usr/bin/env python3
"""Acquire source-native BPBD disaster context datasets from Satu Data Sumbar.

This complements the validated 2024 district-impact lane. It deliberately keeps
2023 losses/continuity data and 2024 mitigation/monthly/type data source-native
until their worksheet semantics and totals are reviewed. No missing values are
inferred and no indicators are merged here.
"""

from __future__ import annotations

import csv
import hashlib
import json
import re
import urllib.parse
import urllib.request
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
OUTPUT_ROOT = ROOT / "data/processed/bpbd/sumbar_open_data"
MANIFEST = ROOT / "data/manifests/sumbar_bpbd_priority_context.json"

CKAN_BASES = (
    "https://data.sumbarprov.go.id/api/3/action",
    "https://data.sumbarprov.go.id/id/api/3/action",
)

PACKAGE_SPECS = (
    {
        "slug": "data-kerugian-akiban-bencana-tahun-2023",
        "title": "Data Kerugian Akibat Bencana Tahun 2023",
        "year": 2023,
        "role": "economic_loss",
    },
    {
        "slug": "dampak-bencana-terhadap-masyarakat-per-kabupaten-kota",
        "title": "Dampak bencana terhadap masyarakat Per Kabupaten/Kota 2023",
        "year": 2023,
        "role": "impact_continuity",
    },
    {
        "slug": "dampak-bencana-terhadap-pemukiman-masyarakat-per-kabupaten-kota-tahun-2023",
        "title": "Dampak Bencana Terhadap Pemukiman Masyarakat Per Kabupaten/Kota Tahun 2023",
        "year": 2023,
        "role": "housing_continuity",
    },
    {
        "slug": "jumlah-korban-perjenis-bencana-2024",
        "title": "Jumlah Korban Perjenis Bencana 2024",
        "year": 2024,
        "role": "casualties_by_hazard",
    },
    {
        "slug": "jumlah-kejadian-bencana-perbulan-2024",
        "title": "Jumlah Kejadian Bencana Perbulan 2024",
        "year": 2024,
        "role": "monthly_events",
    },
    {
        "slug": "jumlah-sirine-tsunami-milik-provinsi-sumatera-barat",
        "title": "Jumlah Sirine Tsunami Milik Provinsi Sumatera Barat",
        "year": 2024,
        "role": "mitigation_capacity",
    },
)


def fetch_bytes(url: str, timeout: int = 120) -> bytes:
    request = urllib.request.Request(
        url,
        headers={"User-Agent": "RanahObservatory/1.0 (+https://github.com/nabilrn/ranah-observatory)"},
    )
    with urllib.request.urlopen(request, timeout=timeout) as response:
        return response.read()


def fetch_json(url: str, timeout: int = 60) -> dict:
    return json.loads(fetch_bytes(url, timeout=timeout).decode("utf-8"))


def ckan_action(action: str, **params: str) -> dict:
    query = urllib.parse.urlencode(params)
    errors: list[str] = []
    for base in CKAN_BASES:
        url = f"{base}/{action}?{query}"
        try:
            payload = fetch_json(url)
            if payload.get("success") is not True:
                raise RuntimeError(f"success=false: {payload}")
            return payload["result"]
        except Exception as exc:
            errors.append(f"{url}: {exc}")
    raise RuntimeError("all official CKAN endpoints failed:\n" + "\n".join(errors))


def resolve_package(spec: dict) -> tuple[str, dict, str]:
    slug = spec["slug"]
    try:
        return slug, ckan_action("package_show", id=slug), "package_show"
    except Exception:
        result = ckan_action("package_search", q=spec["title"], rows="20")
        exact = [
            item for item in result.get("results", [])
            if str(item.get("title", "")).strip().casefold() == spec["title"].casefold()
            and str((item.get("organization") or {}).get("name", "")) == "badan-penanggulangan-bencana-daerah"
        ]
        if len(exact) != 1:
            raise RuntimeError(
                f"could not uniquely resolve {spec['title']!r}; exact BPBD matches={len(exact)}"
            )
        package = exact[0]
        return str(package["name"]), package, "exact_title_search"


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def sha256_path(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def safe_component(value: str) -> str:
    return re.sub(r"[^a-zA-Z0-9._-]+", "-", value.strip()).strip("-") or "sheet"


def cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value)


def package_extras(package: dict) -> dict[str, str]:
    result: dict[str, str] = {}
    for item in package.get("extras", []) or []:
        key = str(item.get("key", "")).strip()
        value = str(item.get("value", "")).strip()
        if key:
            result[key] = value
    return result


def materialize(spec: dict) -> dict:
    resolved_slug, package, resolution = resolve_package(spec)
    resources = [
        item for item in package.get("resources", [])
        if str(item.get("format", "")).strip().upper() == "XLSX"
        and str(item.get("url", "")).startswith("https://data.sumbarprov.go.id/")
    ]
    if not resources:
        raise RuntimeError(f"{resolved_slug}: no official XLSX resource found")

    resources.sort(key=lambda item: (str(item.get("name", "")).casefold(), str(item.get("id", ""))))
    resource = resources[0]
    workbook_bytes = fetch_bytes(str(resource["url"]))
    workbook = load_workbook(BytesIO(workbook_bytes), read_only=True, data_only=True)

    package_dir = OUTPUT_ROOT / str(spec["year"]) / resolved_slug
    package_dir.mkdir(parents=True, exist_ok=True)
    for old_csv in package_dir.glob("*.csv"):
        old_csv.unlink()

    sheets: list[dict] = []
    for index, worksheet in enumerate(workbook.worksheets, start=1):
        output = package_dir / f"{index:02d}-{safe_component(worksheet.title)}.csv"
        row_count = 0
        nonempty_rows = 0
        max_columns = 0
        with output.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            for row in worksheet.iter_rows(values_only=True):
                values = [cell_text(value) for value in row]
                while values and values[-1] == "":
                    values.pop()
                writer.writerow(values)
                row_count += 1
                max_columns = max(max_columns, len(values))
                nonempty_rows += int(any(value != "" for value in values))
        sheets.append(
            {
                "sheet_index": index,
                "sheet_title": worksheet.title,
                "path": output.relative_to(ROOT).as_posix(),
                "sha256": sha256_path(output),
                "row_count": row_count,
                "nonempty_rows": nonempty_rows,
                "max_columns": max_columns,
            }
        )

    extras = package_extras(package)
    return {
        "requested_slug": spec["slug"],
        "resolved_slug": resolved_slug,
        "resolution": resolution,
        "role": spec["role"],
        "year": spec["year"],
        "package_id": package.get("id"),
        "title": package.get("title"),
        "organization": (package.get("organization") or {}).get("title"),
        "producer": extras.get("Produsen Data") or package.get("author"),
        "source_data": extras.get("Sumber Data"),
        "metadata_created": package.get("metadata_created"),
        "metadata_modified": package.get("metadata_modified"),
        "resource_id": resource.get("id"),
        "resource_name": resource.get("name"),
        "resource_format": resource.get("format"),
        "resource_url": resource.get("url"),
        "resource_last_modified": resource.get("last_modified"),
        "download_sha256": sha256_bytes(workbook_bytes),
        "download_size_bytes": len(workbook_bytes),
        "worksheets": sheets,
    }


def main() -> None:
    retrieved_at = datetime.now(timezone.utc).isoformat()
    records: list[dict] = []
    for spec in PACKAGE_SPECS:
        record = materialize(spec)
        records.append(record)
        print(json.dumps({
            "role": record["role"],
            "year": record["year"],
            "package": record["resolved_slug"],
            "resolution": record["resolution"],
            "worksheets": len(record["worksheets"]),
            "sha256": record["download_sha256"],
        }, ensure_ascii=False))

    MANIFEST.parent.mkdir(parents=True, exist_ok=True)
    MANIFEST.write_text(
        json.dumps(
            {
                "schema": "ranah-observatory/sumbar-bpbd-priority-context-acquisition/v1",
                "retrieved_at": retrieved_at,
                "source": "https://data.sumbarprov.go.id/",
                "organization": "BPBD Provinsi Sumatera Barat",
                "promotion_state": "source_native_review_required",
                "missing_values_inferred": False,
                "package_count": len(records),
                "packages": records,
            },
            ensure_ascii=False,
            indent=2,
        ) + "\n",
        encoding="utf-8",
    )
    print(json.dumps({"manifest": MANIFEST.relative_to(ROOT).as_posix(), "package_count": len(records)}))


if __name__ == "__main__":
    main()
