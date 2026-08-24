#!/usr/bin/env python3
"""Build M42 source-native BNPB historical staging from local official XLSX files.

This builder deliberately stops before a canonical longitudinal panel. Historical
raw codes are preserved as source identity and canonical entity IDs are resolved
by exact source-name identity plus the M41 legal-lineage contract, never through
BNPB's 2024 raw-code crosswalk.
"""

from __future__ import annotations

import argparse
import csv
import gzip
import hashlib
import io
import json
import re
from pathlib import Path


YEARS = range(2000, 2018)
METRICS = [
    "jumlah_kejadian",
    "meninggal",
    "hilang",
    "terluka",
    "menderita",
    "mengungsi",
    "rumah_rusak_berat",
    "rumah_rusak_sedang",
    "rumah_rusak_ringan",
    "rumah_terendam",
    "fasilitas_pendidikan",
    "fasilitas_kesehatan",
    "fasilitas_peribadatan",
    "fasilitas_umum",
]
EXPECTED_SOURCE_COLUMNS = [
    "No",
    "Wilayah",
    "Jumlah Kejadian",
    "Meninggal",
    "Hilang",
    "Terluka",
    "Menderita",
    "Mengungsi",
    "Rusak Berat",
    "Rusak Sedang",
    "Rusak Ringan",
    "Terendam",
    "Pendidikan",
    "Kesehatan",
    "Peribadatan",
    "Umum",
]
BASE_FIELDS = [
    "source_year",
    "source_file_sha256",
    "source_sheet",
    "source_row_number",
    "source_label_raw",
    "source_code_raw",
    "source_name_raw",
    "canonical_entity_id_by_name_lineage",
    "geography_lineage_status",
    "source_label_timing_status",
    "current_boundary_comparability",
]
FIELDS = BASE_FIELDS + [
    field
    for metric in METRICS
    for field in (f"{metric}_raw", f"{metric}_value", f"{metric}_state")
]
LABEL_RE = re.compile(r"^\s*(\d+)\.\s*(.+?)\s*$")
COUNT_RE = re.compile(r"^\d+$")


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(65536), b""):
            digest.update(chunk)
    return digest.hexdigest()


def parse_count(value):
    if value is None or (isinstance(value, str) and not value.strip()):
        return None, "source_blank"
    if isinstance(value, bool):
        return None, "source_non_numeric"
    if isinstance(value, int):
        number = value
    elif isinstance(value, float):
        if not value.is_integer():
            return None, "source_non_numeric"
        number = int(value)
    else:
        normalized = str(value).strip().replace(",", "")
        if not COUNT_RE.fullmatch(normalized):
            return None, "source_non_numeric"
        number = int(normalized)
    if number < 0:
        return None, "source_non_numeric"
    return number, "observed_zero" if number == 0 else "observed_positive"


def parse_source_label(value: str) -> tuple[str, str]:
    match = LABEL_RE.fullmatch(str(value))
    if not match:
        raise ValueError(f"unparseable historical Wilayah label: {value!r}")
    return match.group(1), match.group(2)


def load_name_identity_map(path: Path) -> dict[str, str]:
    """Use only exact source-name identity from the 2024 registry.

    The current source-code column is intentionally ignored because M41 proved
    that BNPB historical raw codes follow a different regime.
    """
    result: dict[str, str] = {}
    with path.open(encoding="utf-8", newline="") as handle:
        for row in csv.DictReader(handle):
            name = row["source_name_expected"].strip().upper()
            canonical = row["canonical_geography_id"].strip()
            if name in result and result[name] != canonical:
                raise ValueError(f"ambiguous canonical name identity: {name}")
            result[name] = canonical
    return result


def geography_flags(name: str, year: int) -> tuple[str, str]:
    if name == "SOLOK":
        if year < 2003:
            lineage = "pre_solok_selatan_split_parent"
        elif year == 2003:
            lineage = "transition_year_parent_solok_selatan_split"
        else:
            lineage = "post_solok_selatan_split_lineage"
    elif name == "SIJUNJUNG":
        if year < 2003:
            lineage = "pre_dharmasraya_split_parent"
        elif year == 2003:
            lineage = "transition_year_parent_dharmasraya_split"
        else:
            lineage = "post_dharmasraya_split_lineage"
    elif name == "PASAMAN":
        if year < 2003:
            lineage = "pre_pasaman_barat_split_parent"
        elif year == 2003:
            lineage = "transition_year_parent_pasaman_barat_split"
        else:
            lineage = "post_pasaman_barat_split_lineage"
    elif name == "PADANG PARIAMAN":
        if year < 2002:
            lineage = "pre_kota_pariaman_split_parent"
        elif year == 2002:
            lineage = "transition_year_parent_kota_pariaman_split"
        else:
            lineage = "post_kota_pariaman_split_lineage"
    elif name == "KOTA PARIAMAN":
        if year < 2002:
            lineage = "entity_not_active"
        elif year == 2002:
            lineage = "partial_year_creation"
        else:
            lineage = "post_creation_lineage"
    elif name in {"SOLOK SELATAN", "DHARMASRAYA", "PASAMAN BARAT"}:
        if year < 2003:
            lineage = "entity_not_active"
        elif year == 2003:
            lineage = "partial_year_creation"
        else:
            lineage = "post_creation_lineage"
    elif name == "KEPULAUAN MENTAWAI":
        lineage = "post_1999_creation_lineage"
    else:
        lineage = "no_frozen_boundary_transition_in_m41"

    if name == "SIJUNJUNG" and year < 2008:
        label_status = "retrospective_or_source_normalized_name_before_legal_rename"
    elif name == "SIJUNJUNG" and year == 2008:
        label_status = "legal_rename_transition_year"
    else:
        label_status = "no_known_label_timing_conflict_in_m41"
    return lineage, label_status


def structural_schema(ws) -> list[str]:
    return [
        ws.cell(5, 1).value,
        ws.cell(5, 2).value,
        ws.cell(5, 3).value,
        *[ws.cell(7, column).value for column in range(4, 17)],
    ]


def parse_workbook(path: Path, year: int, canonical_by_name: dict[str, str]):
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - operational dependency guard
        raise RuntimeError("openpyxl is required to parse source workbooks") from exc

    workbook = openpyxl.load_workbook(path, data_only=True)
    if "statistik" not in workbook.sheetnames:
        raise ValueError(f"{path.name}: missing statistik sheet")
    ws = workbook["statistik"]
    if structural_schema(ws) != EXPECTED_SOURCE_COLUMNS:
        raise ValueError(f"{path.name}: structural schema drift")
    expected_label = f"Propinsi : 13. Sumatera Barat, {year}"
    if ws["B4"].value != expected_label:
        raise ValueError(f"{path.name}: unexpected province/year label {ws['B4'].value!r}")

    file_digest = sha256_file(path)
    output_rows = []
    parsed_body = []
    source_total = None

    for row_number in range(9, ws.max_row + 1):
        source_label = ws.cell(row_number, 2).value
        if source_label is None:
            continue
        raw_metrics = [ws.cell(row_number, column).value for column in range(3, 17)]
        if str(source_label).strip().lower() == "jumlah":
            source_total = raw_metrics
            continue

        source_code, source_name = parse_source_label(str(source_label))
        if source_name not in canonical_by_name:
            raise ValueError(f"{path.name}: unmapped source name {source_name!r}")
        lineage, label_timing = geography_flags(source_name, year)
        record = {
            "source_year": year,
            "source_file_sha256": file_digest,
            "source_sheet": "statistik",
            "source_row_number": row_number,
            "source_label_raw": str(source_label),
            "source_code_raw": source_code,
            "source_name_raw": source_name,
            "canonical_entity_id_by_name_lineage": canonical_by_name[source_name],
            "geography_lineage_status": lineage,
            "source_label_timing_status": label_timing,
            "current_boundary_comparability": "not_proven",
        }
        parsed_values = []
        for metric, raw in zip(METRICS, raw_metrics):
            parsed, state = parse_count(raw)
            record[f"{metric}_raw"] = "" if raw is None else str(raw)
            record[f"{metric}_value"] = "" if parsed is None else parsed
            record[f"{metric}_state"] = state
            parsed_values.append(parsed)
        output_rows.append(record)
        parsed_body.append(parsed_values)

    states = [row[f"{metric}_state"] for row in output_rows for metric in METRICS]
    if any(state not in {"observed_zero", "observed_positive"} for state in states):
        raise ValueError(f"{path.name}: blank/non-numeric metric cell blocks M42 staging")

    total_reconciles = None
    if source_total is not None:
        parsed_total = [parse_count(value)[0] for value in source_total]
        sums = [sum(row[index] for row in parsed_body) for index in range(14)]
        total_reconciles = parsed_total == sums
        if not total_reconciles:
            raise ValueError(f"{path.name}: source total does not reconcile across 14 metrics")

    if year == 2001:
        if output_rows or source_total is not None:
            raise ValueError("2001 must remain the frozen empty_body workbook")
        state = "empty_body"
    else:
        if not output_rows or source_total is None:
            raise ValueError(f"{path.name}: expected observed body plus source total")
        state = "observed_body"

    qa = {
        "year": year,
        "sha256": file_digest,
        "bytes": path.stat().st_size,
        "range": f"A1:P{ws.max_row}",
        "rows": len(output_rows),
        "state": state,
        "total": source_total is not None,
        "reconcile14": total_reconciles,
        "cells": len(states),
        "zero_cells": states.count("observed_zero"),
        "positive_cells": states.count("observed_positive"),
        "blank_cells": states.count("source_blank"),
        "nonnumeric_cells": states.count("source_non_numeric"),
    }
    return output_rows, qa


def build(input_dir: Path, geography_map: Path):
    canonical_by_name = load_name_identity_map(geography_map)
    all_rows = []
    qa = []
    for year in YEARS:
        path = input_dir / f"stat_by_wil_13_{year}.xlsx"
        if not path.is_file():
            raise FileNotFoundError(path)
        rows, workbook_qa = parse_workbook(path, year, canonical_by_name)
        all_rows.extend(rows)
        qa.append(workbook_qa)
    return all_rows, qa


def render_csv(rows: list[dict]) -> bytes:
    text = io.StringIO()
    writer = csv.DictWriter(text, fieldnames=FIELDS, lineterminator="\n")
    writer.writeheader()
    writer.writerows(rows)
    return text.getvalue().encode("utf-8")


def deterministic_gzip(payload: bytes) -> bytes:
    target = io.BytesIO()
    with gzip.GzipFile(fileobj=target, mode="wb", filename="", mtime=0) as handle:
        handle.write(payload)
    return target.getvalue()


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input-dir", type=Path, required=True)
    parser.add_argument(
        "--geography-map",
        type=Path,
        default=Path("data/registries/bnpb_geography_map.csv"),
    )
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--qa-json", type=Path)
    args = parser.parse_args()

    rows, qa = build(args.input_dir, args.geography_map)
    csv_bytes = render_csv(rows)
    compressed = deterministic_gzip(csv_bytes)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_bytes(compressed)

    report = {
        "row_count": len(rows),
        "metric_cell_count": len(rows) * len(METRICS),
        "uncompressed_sha256": hashlib.sha256(csv_bytes).hexdigest(),
        "compressed_sha256": hashlib.sha256(compressed).hexdigest(),
        "workbooks": qa,
        "canonical_panel_authorized": False,
    }
    if args.qa_json:
        args.qa_json.parent.mkdir(parents=True, exist_ok=True)
        args.qa_json.write_text(json.dumps(report, indent=2) + "\n", encoding="utf-8")
    else:
        print(json.dumps(report, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
