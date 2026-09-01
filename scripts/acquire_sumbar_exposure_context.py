#!/usr/bin/env python3
"""Acquire machine-readable Sumbar exposure/context workbooks from Satu Data Sumbar.

This lane is source-native only. It captures the official CKAN package/resource
metadata, downloads the XLSX resource, converts every worksheet to CSV, and
records checksums. It does not infer semantics or promote indicators.

Current priority sources support Disaster Explore context:
- SDABK rain-gauge post inventory (local hydro-meteorological monitoring capacity)
- Provincial-road length by surface type (infrastructure exposure/context)
"""

from __future__ import annotations

import csv
import hashlib
import json
import re
import urllib.parse
import urllib.request
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
OUTPUT_ROOT = ROOT / "data/processed/sumbarprov/exposure_context"
MANIFEST = ROOT / "data/manifests/sumbar_exposure_context.json"

CKAN_BASES = (
    "https://data.sumbarprov.go.id/api/3/action",
    "https://data.sumbarprov.go.id/id/api/3/action",
)

PACKAGE_SPECS = (
    {
        "slug": "jumlah-pos-curah-hujan-dinas-sdabk-tahun-2024",
        "title": "Jumlah Pos Curah Hujan Dinas SDABK Tahun 2024",
        "organization": "dinas-pengelolaan-sumber-daya-air-dan-bina-konstruksi",
        "role": "rain_gauge_inventory",
        "domain": "climate_monitoring_context",
    },
    {
        "slug": "panjang-jalan-provinsi-berdasarkan-jenis-permukaan-km",
        "title": "Panjang Jalan Provinsi Berdasarkan Jenis Permukaan (Km)",
        "organization": "dinas-bina-marga-cipta-karya-dan-tata-ruang",
        "role": "provincial_road_surface_length",
        "domain": "infrastructure_exposure_context",
    },
)


def fetch_bytes(url: str, timeout: int = 120) -> bytes:
    request = urllib.request.Request(
        url,
        headers={"User-Agent": "RanahObservatory/1.0 (+https://github.com/nabilrn/ranah-observatory)"},
    )
    with urllib.request.urlopen(request, timeout=timeout) as response:
        return response.read()


def fetch_json(url: str, timeout: int = 60) -> dict:
    return json.loads(fetch_bytes(url, timeout=timeout).decode("utf-8"))


def ckan_action(action: str, **params: str) -> dict:
    query = urllib.parse.urlencode(params)
    errors: list[str] = []
    for base in CKAN_BASES:
        url = f"{base}/{action}?{query}"
        try:
            payload = fetch_json(url)
            if payload.get("success") is not True:
                raise RuntimeError(f"success=false: {payload}")
            return payload["result"]
        except Exception as exc:
            errors.append(f"{url}: {exc}")
    raise RuntimeError("all official CKAN endpoints failed:\n" + "\n".join(errors))


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def sha256_path(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def safe_component(value: str) -> str:
    return re.sub(r"[^a-zA-Z0-9._-]+", "-", value.strip()).strip("-") or "sheet"


def cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value)


def package_extras(package: dict) -> dict[str, str]:
    result: dict[str, str] = {}
    for item in package.get("extras", []) or []:
        key = str(item.get("key", "")).strip()
        value = str(item.get("value", "")).strip()
        if key:
            result[key] = value
    return result


def materialize(spec: dict) -> dict:
    package = ckan_action("package_show", id=spec["slug"])
    if str(package.get("title", "")).strip() != spec["title"]:
        raise RuntimeError(
            f"{spec['slug']}: title changed: expected={spec['title']!r} actual={package.get('title')!r}"
        )
    organization = package.get("organization") or {}
    if str(organization.get("name", "")).strip() != spec["organization"]:
        raise RuntimeError(
            f"{spec['slug']}: organization changed: expected={spec['organization']!r} actual={organization.get('name')!r}"
        )

    resources = [
        item for item in package.get("resources", [])
        if str(item.get("format", "")).strip().upper() == "XLSX"
        and str(item.get("url", "")).startswith("https://data.sumbarprov.go.id/")
    ]
    if len(resources) != 1:
        raise RuntimeError(f"{spec['slug']}: expected exactly one official XLSX resource, found {len(resources)}")
    resource = resources[0]
    workbook_bytes = fetch_bytes(str(resource["url"]))
    workbook = load_workbook(BytesIO(workbook_bytes), read_only=True, data_only=True)

    package_dir = OUTPUT_ROOT / spec["slug"]
    package_dir.mkdir(parents=True, exist_ok=True)
    for old_csv in package_dir.glob("*.csv"):
        old_csv.unlink()

    sheets: list[dict] = []
    for index, worksheet in enumerate(workbook.worksheets, start=1):
        output = package_dir / f"{index:02d}-{safe_component(worksheet.title)}.csv"
        row_count = 0
        nonempty_rows = 0
        max_columns = 0
        with output.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            for row in worksheet.iter_rows(values_only=True):
                values = [cell_text(value) for value in row]
                while values and values[-1] == "":
                    values.pop()
                writer.writerow(values)
                row_count += 1
                max_columns = max(max_columns, len(values))
                nonempty_rows += int(any(value != "" for value in values))
        sheets.append(
            {
                "sheet_index": index,
                "sheet_title": worksheet.title,
                "path": output.relative_to(ROOT).as_posix(),
                "sha256": sha256_path(output),
                "row_count": row_count,
                "nonempty_rows": nonempty_rows,
                "max_columns": max_columns,
            }
        )

    extras = package_extras(package)
    return {
        "slug": spec["slug"],
        "title": package.get("title"),
        "role": spec["role"],
        "domain": spec["domain"],
        "package_id": package.get("id"),
        "organization_name": organization.get("name"),
        "organization_title": organization.get("title"),
        "author": package.get("author"),
        "producer": extras.get("Produsen Data"),
        "source_data": extras.get("Sumber Data"),
        "metadata_created": package.get("metadata_created"),
        "metadata_modified": package.get("metadata_modified"),
        "tags": sorted(str(item.get("name", "")).strip() for item in package.get("tags", []) if item.get("name")),
        "resource_id": resource.get("id"),
        "resource_name": resource.get("name"),
        "resource_format": resource.get("format"),
        "resource_url": resource.get("url"),
        "resource_last_modified": resource.get("last_modified"),
        "download_sha256": sha256_bytes(workbook_bytes),
        "download_size_bytes": len(workbook_bytes),
        "worksheets": sheets,
    }


def main() -> None:
    retrieved_at = datetime.now(timezone.utc).isoformat()
    packages: list[dict] = []
    for spec in PACKAGE_SPECS:
        record = materialize(spec)
        packages.append(record)
        print(json.dumps({
            "role": record["role"],
            "package": record["slug"],
            "organization": record["organization_title"],
            "worksheets": len(record["worksheets"]),
            "sha256": record["download_sha256"],
        }, ensure_ascii=False))

    MANIFEST.parent.mkdir(parents=True, exist_ok=True)
    MANIFEST.write_text(
        json.dumps(
            {
                "schema": "ranah-observatory/sumbar-exposure-context-acquisition/v1",
                "retrieved_at": retrieved_at,
                "source": "https://data.sumbarprov.go.id/",
                "promotion_state": "source_native_review_required",
                "missing_values_inferred": False,
                "package_count": len(packages),
                "packages": packages,
            },
            ensure_ascii=False,
            indent=2,
        ) + "\n",
        encoding="utf-8",
    )
    print(json.dumps({"manifest": MANIFEST.relative_to(ROOT).as_posix(), "package_count": len(packages)}))


if __name__ == "__main__":
    main()
